"""Génère un formulaire Kobo (XLSForm) dont les noms de questions correspondent
EXACTEMENT aux identifiants du schéma de l'étoile du conseil — c'est ce qui
permet au connecteur (modules/kobo_import.py) de remapper chaque réponse
directement dans la bonne branche/champ, sans ambiguïté ni intervention IA.

Utilisation : l'administrateur télécharge ce fichier .xlsx et l'importe tel
quel dans KoboToolbox ("New" > "Import an XLSForm") pour déployer le
formulaire de collecte terrain.
"""
import openpyxl

from modules.collecte import load_schema

_TYPE_MAP_NUMBER = "decimal"


def _slug_options(options: list) -> list:
    """Attribue un nom de code stable (opt1, opt2...) à chaque option, pour
    éviter tout souci d'encodage/accents dans les noms de choix Kobo — le
    libellé affiché, lui, reste le texte complet en FR/EN."""
    return [f"opt{i + 1}" for i in range(len(options))]


def generate_kobo_xlsform() -> openpyxl.Workbook:
    schema = load_schema()
    wb = openpyxl.Workbook()

    survey_ws = wb.active
    survey_ws.title = "survey"
    survey_ws.append(["type", "name", "label::French (fr)", "label::English (en)", "appearance", "required"])

    choices_ws = wb.create_sheet("choices")
    choices_ws.append(["list_name", "name", "label::French (fr)", "label::English (en)"])

    settings_ws = wb.create_sheet("settings")
    settings_ws.append(["form_title", "form_id"])
    settings_ws.append(["Diagnostic Conseil Agropastoral", "diagnostic_conseil_agropastoral"])

    for branch_key, branch in schema["branches"].items():
        label_fr = branch["label"]["fr"]
        label_en = branch["label"].get("en", label_fr)
        survey_ws.append(["begin group", branch_key, label_fr, label_en, "field-list", ""])

        for field in branch["fields"]:
            fid = field["id"]
            flabel_fr = field["label"]["fr"]
            flabel_en = field["label"].get("en", flabel_fr)
            ftype = field["type"]

            if ftype == "number":
                survey_ws.append([_TYPE_MAP_NUMBER, fid, flabel_fr, flabel_en, "", ""])
            elif ftype == "textarea":
                survey_ws.append(["text", fid, flabel_fr, flabel_en, "multiline", ""])
            elif ftype == "select":
                list_name = f"{branch_key}_{fid}"
                options_fr = field["options"]["fr"]
                options_en = field["options"].get("en", options_fr)
                codes = _slug_options(options_fr)
                for code, opt_fr, opt_en in zip(codes, options_fr, options_en):
                    choices_ws.append([list_name, code, opt_fr, opt_en])
                survey_ws.append([f"select_one {list_name}", fid, flabel_fr, flabel_en, "", ""])
            elif ftype == "activity_list":
                survey_ws.append(["begin repeat", fid, flabel_fr, flabel_en, "", ""])
                survey_ws.append(["text", "nom", "Nom de l'activité/produit", "Activity/product name", "", ""])
                survey_ws.append([_TYPE_MAP_NUMBER, "part_marche_relative",
                                  "Part de marché relative (0 à 5)", "Relative market share (0 to 5)", "", ""])
                survey_ws.append([_TYPE_MAP_NUMBER, "taux_croissance",
                                  "Taux de croissance du marché (%)", "Market growth rate (%)", "", ""])
                survey_ws.append(["end repeat", "", "", "", "", ""])
            else:
                survey_ws.append(["text", fid, flabel_fr, flabel_en, "", ""])

        survey_ws.append(["end group", "", "", "", "", ""])

    return wb


def generate_kobo_xlsform_bytes() -> bytes:
    import io
    wb = generate_kobo_xlsform()
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
